import pytest
from openpyxl.cell import Cell

from excel_db.columns import Column
from excel_db.db import ExcelDB
from excel_db.models import ExcelModel


class User(ExcelModel):
    id = Column()
    name = Column()


class MyDB(ExcelDB):
    users = User.as_table()


@pytest.fixture()
def db(tmp_excel_file):
    return MyDB(tmp_excel_file)


def test_table_cell(db):
    cell = db.users.cell(1, 1)
    assert isinstance(cell, Cell)
    assert cell.row == 1
    assert cell.column == 1


def test_row_cell(db):
    cell = db.users[1].cell(2)
    assert isinstance(cell, Cell)
    assert cell.row == 3
    assert cell.column == 2


def test_row_cell0(db):
    cell = db.users[2].cell0(0)
    assert isinstance(cell, Cell)
    assert cell.row == 4
    assert cell.column == 1


def test_row_cella(db):
    cell = db.users[0].cella('id')
    assert isinstance(cell, Cell)
    assert cell.row == 2
    assert cell.column == 1


def test_col_cell(db):
    cell = db.users.id.cell(3)
    assert isinstance(cell, Cell)
    assert cell.row == 3
    assert cell.column == 1


def test_col_cell0(db):
    cell = db.users.name.cell0(0)
    assert isinstance(cell, Cell)
    assert cell.row == 2
    assert cell.column == 2
